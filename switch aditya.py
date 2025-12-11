import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import customtkinter as ctk
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import threading
import time
import os
from datetime import datetime
import calendar

class LoadingWindow(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Processing")
        
        # Configure window
        self.geometry("300x150")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        
        # Center the window
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')
        
        # Create loading frame
        self.loading_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.loading_frame.pack(expand=True, fill="both", padx=20, pady=20)
        
        # Loading label
        self.loading_label = ctk.CTkLabel(
            self.loading_frame,
            text="Processing your files...",
            font=("Segoe UI", 14, "bold"),
            text_color="#ffffff"
        )
        self.loading_label.pack(pady=(0, 15))
        
        # Progress bar
        self.progress_bar = ctk.CTkProgressBar(self.loading_frame)
        self.progress_bar.pack(fill="x", padx=20)
        self.progress_bar.set(0)
        
        # Status label
        self.status_label = ctk.CTkLabel(
            self.loading_frame,
            text="Initializing...",
            font=("Segoe UI", 12),
            text_color="#b0b0b0"
        )
        self.status_label.pack(pady=(10, 0))
        
        # Start animation
        self.animation_running = True
        self.animate()

    def animate(self):
        if not self.animation_running:
            return
        
        # Update progress bar
        current = self.progress_bar.get()
        if current >= 1:
            self.progress_bar.set(0)
        else:
            self.progress_bar.set(current + 0.1)
        
        # Schedule next update
        self.after(100, self.animate)

    def update_status(self, status):
        self.status_label.configure(text=status)

    def stop(self):
        self.animation_running = False
        self.destroy()

class SwitchRegisterGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Switch Register - Launcher")
        self.root.geometry("800x700")
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")
        self.root.resizable(False, False)
        
        # Store selected file paths
        self.switch_register_path = None
        self.rta_master_path = None
        self.brokerage_structure_paths = []  # List to store multiple files
        
        # Create UI
        self.create_widgets()
        
        # Center the window
        self.center_window()
    
    def center_window(self):
        """Center the window on screen"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def create_widgets(self):
        """Create and arrange all GUI widgets"""
        # Main container with dark background
        main_frame = ctk.CTkFrame(self.root, fg_color="#1a1a1a")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Title section
        title_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        title_frame.pack(pady=(20, 10))
        
        title_label = ctk.CTkLabel(
            title_frame,
            text="SWITCH REGISTER",
            font=("Segoe UI", 32, "bold"),
            text_color="#ffffff"
        )
        title_label.pack()
        
        subtitle_label = ctk.CTkLabel(
            title_frame,
            text="Upload Center",
            font=("Segoe UI", 16),
            text_color="#b0b0b0"
        )
        subtitle_label.pack(pady=(5, 0))
        
        # Central panel with rounded corners
        upload_panel = ctk.CTkFrame(main_frame, fg_color="#2b2b2b", corner_radius=15)
        upload_panel.pack(fill="both", expand=True, padx=30, pady=20)
        
        # Instruction text
        instruction_label = ctk.CTkLabel(
            upload_panel,
            text="Select files to begin",
            font=("Segoe UI", 14),
            text_color="#b0b0b0"
        )
        instruction_label.pack(pady=(30, 30))
        
        # File upload sections
        # 1. Switch Register (Blue)
        self.create_upload_section(
            upload_panel,
            "SWITCH REGISTER",
            "#3498db",
            "#2980b9",
            self.upload_switch_register,
            "switch_register"
        )
        
        # 2. RTA Master (Light Blue/Cyan)
        self.create_upload_section(
            upload_panel,
            "RTA MASTER",
            "#5dade2",
            "#3498db",
            self.upload_rta_master,
            "rta_master"
        )
        
        # 3. Brokerage Structure File (Green) - Multiple files
        self.create_upload_section(
            upload_panel,
            "BROKERAGE STRUCTURE FILE",
            "#27ae60",
            "#219a52",
            self.upload_brokerage_structure,
            "brokerage_structure",
            multiple=True
        )
        
        # Process button at bottom
        process_btn = ctk.CTkButton(
            upload_panel,
            text="PROCESS",
            command=self.process_files,
            fg_color="#27ae60",
            hover_color="#219a52",
            font=("Segoe UI", 16, "bold"),
            height=50,
            corner_radius=10
        )
        process_btn.pack(fill="x", padx=30, pady=(20, 30))
        
        # Status bar
        self.status_label = ctk.CTkLabel(
            main_frame,
            text="Ready to upload files",
            font=("Segoe UI", 11),
            text_color="#b0b0b0"
        )
        self.status_label.pack(pady=(0, 10))
    
    def create_upload_section(self, parent, label_text, button_color, hover_color, command, file_type, multiple=False):
        """Create a file upload section with button and status label"""
        section_frame = ctk.CTkFrame(parent, fg_color="transparent")
        section_frame.pack(fill="x", padx=30, pady=10)
        
        # Upload button
        upload_btn = ctk.CTkButton(
            section_frame,
            text=label_text,
            command=command,
            fg_color=button_color,
            hover_color=hover_color,
            font=("Segoe UI", 12, "bold"),
            width=200,
            height=40,
            corner_radius=8
        )
        upload_btn.pack(side="left", padx=(0, 15))
        
        # Status label
        status_label = ctk.CTkLabel(
            section_frame,
            text="No file selected" if not multiple else "No files selected",
            text_color="#b0b0b0",
            font=("Segoe UI", 11),
            anchor="w"
        )
        status_label.pack(side="left", fill="x", expand=True)
        
        # Store reference to status label
        if file_type == "switch_register":
            self.switch_register_status = status_label
        elif file_type == "rta_master":
            self.rta_master_status = status_label
        elif file_type == "brokerage_structure":
            self.brokerage_structure_status = status_label
    
    def upload_switch_register(self):
        """Handle Switch Register file upload"""
        file_path = filedialog.askopenfilename(
            title="Select Switch Register File",
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("CSV files", "*.csv"),
                ("All files", "*.*")
            ]
        )
        if file_path:
            self.switch_register_path = file_path
            filename = os.path.basename(file_path)
            self.switch_register_status.configure(
                text=f"✓ {filename}",
                text_color="#3498db"
            )
            self.update_status()
            messagebox.showinfo("Success", "Switch Register file uploaded successfully!")
    
    def upload_rta_master(self):
        """Handle RTA Master file upload"""
        file_path = filedialog.askopenfilename(
            title="Select RTA Master File",
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("CSV files", "*.csv"),
                ("All files", "*.*")
            ]
        )
        if file_path:
            self.rta_master_path = file_path
            filename = os.path.basename(file_path)
            self.rta_master_status.configure(
                text=f"✓ {filename}",
                text_color="#5dade2"
            )
            self.update_status()
            messagebox.showinfo("Success", "RTA Master file uploaded successfully!")
    
    def upload_brokerage_structure(self):
        """Handle Brokerage Structure file upload (multiple files)"""
        file_paths = filedialog.askopenfilenames(
            title="Select Brokerage Structure File(s)",
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("CSV files", "*.csv"),
                ("All files", "*.*")
            ]
        )
        if file_paths:
            self.brokerage_structure_paths = list(file_paths)
            if len(file_paths) == 1:
                filename = os.path.basename(file_paths[0])
                self.brokerage_structure_status.configure(
                    text=f"✓ {filename}",
                    text_color="#27ae60"
                )
            else:
                self.brokerage_structure_status.configure(
                    text=f"✓ {len(file_paths)} files selected",
                    text_color="#27ae60"
                )
            self.update_status()
            messagebox.showinfo("Success", f"{len(file_paths)} Brokerage Structure file(s) uploaded successfully!")
    
    def update_status(self):
        """Update the main status label"""
        files_uploaded = sum([
            self.switch_register_path is not None,
            self.rta_master_path is not None,
            len(self.brokerage_structure_paths) > 0
        ])
        
        if files_uploaded == 0:
            self.status_label.configure(
                text="Ready to upload files",
                text_color="#b0b0b0"
            )
        elif files_uploaded < 3:
            self.status_label.configure(
                text=f"{files_uploaded} of 3 file types uploaded",
                text_color="#f39c12"
            )
        else:
            self.status_label.configure(
                text="All files uploaded. Ready to process!",
                text_color="#27ae60"
            )
    
    def process_files(self):
        """Handle the process button click"""
        # Check if required files are uploaded
        if not self.switch_register_path:
            messagebox.showwarning(
                "Incomplete Upload",
                "Please upload Switch Register file before processing."
            )
            return
        
        if not self.rta_master_path:
            messagebox.showwarning(
                "Incomplete Upload",
                "Please upload RTA Master file before processing."
            )
            return
        
        if len(self.brokerage_structure_paths) == 0:
            messagebox.showwarning(
                "Incomplete Upload",
                "Please upload at least one Brokerage Structure file before processing."
            )
            return
        
        # Show loading window
        loading_window = LoadingWindow(self.root)
        
        def process_file():
            try:
                loading_window.update_status("Reading Switch Register file...")
                # Read Switch Register file
                if self.switch_register_path.endswith('.csv'):
                    switch_df = pd.read_csv(self.switch_register_path)
                else:
                    switch_df = pd.read_excel(self.switch_register_path)
                
                # Check for duplicate columns
                if switch_df.columns.duplicated().any():
                    dupes = switch_df.columns[switch_df.columns.duplicated()].tolist()
                    self.root.after(0, lambda: messagebox.showerror(
                        "Error",
                        f"Switch Register file has duplicate column names: {dupes}. Please fix the file and try again."
                    ))
                    self.root.after(0, loading_window.stop)
                    return
                
                loading_window.update_status("Reading RTA Master file...")
                # Read RTA Master file
                if self.rta_master_path.endswith('.csv'):
                    rta_df = pd.read_csv(self.rta_master_path)
                else:
                    rta_df = pd.read_excel(self.rta_master_path)
                
                # Check for duplicate columns
                if rta_df.columns.duplicated().any():
                    dupes = rta_df.columns[rta_df.columns.duplicated()].tolist()
                    self.root.after(0, lambda: messagebox.showerror(
                        "Error",
                        f"RTA Master file has duplicate column names: {dupes}. Please fix the file and try again."
                    ))
                    self.root.after(0, loading_window.stop)
                    return
                
                loading_window.update_status("Reading Brokerage Structure file(s)...")
                # Read all Brokerage Structure files
                brokerage_dfs = []
                for file_path in self.brokerage_structure_paths:
                    if file_path.endswith('.csv'):
                        df = pd.read_csv(file_path)
                    else:
                        df = pd.read_excel(file_path)
                    
                    # Check for duplicate columns
                    if df.columns.duplicated().any():
                        dupes = df.columns[df.columns.duplicated()].tolist()
                        self.root.after(0, lambda d=dupes, f=os.path.basename(file_path): messagebox.showerror(
                            "Error",
                            f"Brokerage Structure file '{f}' has duplicate column names: {d}. Please fix the file and try again."
                        ))
                        self.root.after(0, loading_window.stop)
                        return
                    
                    brokerage_dfs.append(df)
                
                # Combine all brokerage structure files
                if len(brokerage_dfs) > 1:
                    combined_brokerage_df = pd.concat(brokerage_dfs, ignore_index=True)
                else:
                    combined_brokerage_df = brokerage_dfs[0] if brokerage_dfs else pd.DataFrame()
                
                loading_window.update_status("Processing Switch Register data...")
                
                # Process Switch Register: Extract scheme codes from "From" and "Scheme :" columns
                processed_df = switch_df.copy()
                
                # Function to extract scheme code (part before "/")
                def extract_scheme_code(value):
                    if pd.isna(value):
                        return None
                    value_str = str(value)
                    if '/' in value_str:
                        scheme_code = value_str.split('/')[0].strip()
                        return scheme_code
                    return None
                
                # Find "From" column (case-insensitive)
                from_col = None
                for col in processed_df.columns:
                    if str(col).upper().strip() == 'FROM':
                        from_col = col
                        break
                
                # Find "Scheme :" column (case-insensitive, handle variations)
                scheme_col = None
                for col in processed_df.columns:
                    col_upper = str(col).upper().strip()
                    if col_upper == 'SCHEME :' or col_upper == 'SCHEME:' or col_upper == 'SCHEME':
                        scheme_col = col
                        break
                
                # Process "From" column
                if from_col:
                    # Extract scheme code and add as "out Scheme Code" at the start
                    out_scheme_codes = processed_df[from_col].apply(extract_scheme_code)
                    processed_df.insert(0, 'out Scheme Code', out_scheme_codes)
                    
                    # Rename "From" column to "switch out scheme"
                    processed_df = processed_df.rename(columns={from_col: 'switch out scheme'})
                else:
                    # If "From" column not found, show warning but continue
                    self.root.after(0, lambda: messagebox.showwarning(
                        "Warning",
                        "From column not found in Switch Register. Processing without out scheme code extraction."
                    ))
                    # Add empty column
                    processed_df.insert(0, 'out Scheme Code', None)
                
                # Process "Scheme :" column
                if scheme_col:
                    # Extract scheme code and add as "IN Scheme Code" after "out Scheme Code"
                    in_scheme_codes = processed_df[scheme_col].apply(extract_scheme_code)
                    # Find position after "out Scheme Code"
                    out_scheme_code_idx = list(processed_df.columns).index('out Scheme Code')
                    processed_df.insert(out_scheme_code_idx + 1, 'IN Scheme Code', in_scheme_codes)
                    
                    # Rename "Scheme :" column to "switch in scheme"
                    processed_df = processed_df.rename(columns={scheme_col: 'switch in scheme'})
                else:
                    # If "Scheme :" column not found, show warning but continue
                    self.root.after(0, lambda: messagebox.showwarning(
                        "Warning",
                        "Scheme : column not found in Switch Register. Processing without IN scheme code extraction."
                    ))
                    # Add empty column after "out Scheme Code"
                    out_scheme_code_idx = list(processed_df.columns).index('out Scheme Code')
                    processed_df.insert(out_scheme_code_idx + 1, 'IN Scheme Code', None)
                
                loading_window.update_status("Matching scheme codes with RTA Master...")
                
                # Find Scheme_code, PARENT_SUB_FUND_CODE, and ASSET_CLASS columns in RTA Master
                scheme_code_col = None
                parent_sub_fund_code_col = None
                asset_class_col = None
                
                for col in rta_df.columns:
                    try:
                        col_upper = str(col).upper().strip()
                        if col_upper == 'SCHEME_CODE' or col_upper == 'SCHEME CODE':
                            scheme_code_col = col
                        elif col_upper == 'PARENT_SUB_FUND_CODE' or col_upper == 'PARENT SUB FUND CODE':
                            parent_sub_fund_code_col = col
                        elif col_upper == 'ASSET_CLASS' or col_upper == 'ASSET CLASS' or ('ASSET' in col_upper and 'CLASS' in col_upper):
                            asset_class_col = col
                    except (TypeError, AttributeError):
                        # Skip columns that can't be converted to string or checked
                        continue
                
                if scheme_code_col and parent_sub_fund_code_col:
                    # Normalize RTA Master columns for matching
                    rta_df_normalized = rta_df.copy()
                    rta_df_normalized[scheme_code_col] = rta_df_normalized[scheme_code_col].astype(str).str.strip()
                    rta_df_normalized[parent_sub_fund_code_col] = rta_df_normalized[parent_sub_fund_code_col].astype(str).str.strip()
                    
                    # Create mapping: Scheme_code -> PARENT_SUB_FUND_CODE
                    subfund_mapping = rta_df_normalized.set_index(scheme_code_col)[parent_sub_fund_code_col].to_dict()
                    
                    # Create mapping: Scheme_code -> ASSET_CLASS (if column exists)
                    asset_class_mapping = {}
                    if asset_class_col:
                        rta_df_normalized[asset_class_col] = rta_df_normalized[asset_class_col].astype(str).str.strip()
                        asset_class_mapping = rta_df_normalized.set_index(scheme_code_col)[asset_class_col].to_dict()
                    
                    # Match "out Scheme Code" with Scheme_code and get PARENT_SUB_FUND_CODE
                    if 'out Scheme Code' in processed_df.columns:
                        processed_df['out Scheme Code'] = processed_df['out Scheme Code'].astype(str).str.strip()
                        processed_df['out subfund code'] = processed_df['out Scheme Code'].map(subfund_mapping)
                        processed_df['out subfund code'] = processed_df['out subfund code'].fillna('Not Found')
                        
                        # Get ASSET_CLASS for "out Scheme Code"
                        if asset_class_mapping:
                            processed_df['out ASSET_CLASS'] = processed_df['out Scheme Code'].map(asset_class_mapping)
                            processed_df['out ASSET_CLASS'] = processed_df['out ASSET_CLASS'].fillna('Not Found')
                        else:
                            processed_df['out ASSET_CLASS'] = 'Not Found'
                        
                        # Insert "out subfund code" right after "out Scheme Code"
                        cols = list(processed_df.columns)
                        out_scheme_idx = cols.index('out Scheme Code')
                        # Remove from current position
                        out_subfund = processed_df.pop('out subfund code')
                        out_asset_class = processed_df.pop('out ASSET_CLASS')
                        # Insert at correct position
                        processed_df.insert(out_scheme_idx + 1, 'out subfund code', out_subfund)
                        processed_df.insert(out_scheme_idx + 2, 'out ASSET_CLASS', out_asset_class)
                    else:
                        processed_df['out subfund code'] = 'Not Found'
                        processed_df['out ASSET_CLASS'] = 'Not Found'
                    
                    # Match "IN Scheme Code" with Scheme_code and get PARENT_SUB_FUND_CODE
                    if 'IN Scheme Code' in processed_df.columns:
                        processed_df['IN Scheme Code'] = processed_df['IN Scheme Code'].astype(str).str.strip()
                        processed_df['IN subfund code'] = processed_df['IN Scheme Code'].map(subfund_mapping)
                        processed_df['IN subfund code'] = processed_df['IN subfund code'].fillna('Not Found')
                        
                        # Get ASSET_CLASS for "IN Scheme Code"
                        if asset_class_mapping:
                            processed_df['IN ASSET_CLASS'] = processed_df['IN Scheme Code'].map(asset_class_mapping)
                            processed_df['IN ASSET_CLASS'] = processed_df['IN ASSET_CLASS'].fillna('Not Found')
                        else:
                            processed_df['IN ASSET_CLASS'] = 'Not Found'
                        
                        # Insert "IN subfund code" right after "IN Scheme Code"
                        cols = list(processed_df.columns)
                        in_scheme_idx = cols.index('IN Scheme Code')
                        # Remove from current position
                        in_subfund = processed_df.pop('IN subfund code')
                        in_asset_class = processed_df.pop('IN ASSET_CLASS')
                        # Insert at correct position
                        processed_df.insert(in_scheme_idx + 1, 'IN subfund code', in_subfund)
                        processed_df.insert(in_scheme_idx + 2, 'IN ASSET_CLASS', in_asset_class)
                    else:
                        processed_df['IN subfund code'] = 'Not Found'
                        processed_df['IN ASSET_CLASS'] = 'Not Found'
                else:
                    # If columns not found, show warning but continue
                    missing_cols = []
                    if not scheme_code_col:
                        missing_cols.append("Scheme_code")
                    if not parent_sub_fund_code_col:
                        missing_cols.append("PARENT_SUB_FUND_CODE")
                    
                    self.root.after(0, lambda: messagebox.showwarning(
                        "Warning",
                        f"Columns not found in RTA Master: {', '.join(missing_cols)}\n"
                        f"Subfund codes and ASSET_CLASS will not be added."
                    ))
                    processed_df['out subfund code'] = 'Not Found'
                    processed_df['out ASSET_CLASS'] = 'Not Found'
                    processed_df['IN subfund code'] = 'Not Found'
                    processed_df['IN ASSET_CLASS'] = 'Not Found'
                
                loading_window.update_status("Matching with Brokerage Structure...")
                
                # Match broker and IN subfund code with Brokerage Structure
                if not combined_brokerage_df.empty:
                    # DEBUG: Print Brokerage Structure columns
                    print("\n=== DEBUG: Brokerage Structure Columns ===")
                    print(list(combined_brokerage_df.columns))
                    print(f"Total rows: {len(combined_brokerage_df)}")
                    if len(combined_brokerage_df) > 0:
                        print("\nFirst few rows:")
                        print(combined_brokerage_df.head(3))
                    
                    # Find columns in Brokerage Structure
                    cons_code_col = None
                    scheme_code_b_col = None
                    trail_rate_cols = {}
                    investment_period_from_col = None
                    investment_period_to_col = None
                    
                    for col in combined_brokerage_df.columns:
                        try:
                            col_upper = str(col).upper().strip()
                            # More flexible matching for Cons Code
                            if 'CONS' in col_upper and 'CODE' in col_upper:
                                cons_code_col = col
                            elif col_upper == 'CONS_CODE' or col_upper == 'CONS':
                                cons_code_col = col
                            # More flexible matching for Scheme Code
                            elif 'SCHEME' in col_upper and 'CODE' in col_upper:
                                scheme_code_b_col = col
                            elif col_upper == 'SCHEME_CODE' or (col_upper == 'SCHEME' and scheme_code_b_col is None):
                                scheme_code_b_col = col
                            # Investment Period From
                            elif 'INVESTMENT' in col_upper and 'PERIOD' in col_upper and 'FROM' in col_upper:
                                investment_period_from_col = col
                            # Investment Period To
                            elif 'INVESTMENT' in col_upper and 'PERIOD' in col_upper and 'TO' in col_upper:
                                investment_period_to_col = col
                            # Trail Rate columns
                            elif 'TRAIL' in col_upper and '1' in col_upper and ('YEAR' in col_upper or 'YR' in col_upper):
                                trail_rate_cols[1] = col
                            elif 'TRAIL' in col_upper and '2' in col_upper and ('YEAR' in col_upper or 'YR' in col_upper):
                                trail_rate_cols[2] = col
                            elif 'TRAIL' in col_upper and '3' in col_upper and ('YEAR' in col_upper or 'YR' in col_upper):
                                trail_rate_cols[3] = col
                            elif 'TRAIL' in col_upper and '4' in col_upper and ('YEAR' in col_upper or 'YR' in col_upper):
                                trail_rate_cols[4] = col
                            elif 'TRAIL' in col_upper and '5' in col_upper and ('YEAR' in col_upper or 'YR' in col_upper):
                                trail_rate_cols[5] = col
                        except (TypeError, AttributeError):
                            # Skip columns that can't be converted to string or checked
                            continue
                    
                    # DEBUG: Print found columns
                    print(f"\n=== DEBUG: Found Columns ===")
                    print(f"Cons Code Column: {cons_code_col}")
                    print(f"Scheme Code Column: {scheme_code_b_col}")
                    print(f"Investment Period From Column: {investment_period_from_col}")
                    print(f"Investment Period To Column: {investment_period_to_col}")
                    print(f"Trail Rate Columns: {trail_rate_cols}")
                    
                    # Find broker column in Switch Register (flexible matching)
                    broker_col = None
                    for col in processed_df.columns:
                        col_upper = str(col).upper().strip()
                        if 'BROK' in col_upper and ('DLR' in col_upper or 'DEALER' in col_upper):
                            broker_col = col
                            break
                        elif col_upper == 'BROKER' or col_upper == 'BROKER CODE' or col_upper == 'BROKER_CODE':
                            broker_col = col
                            break
                    
                    print(f"Broker Column in Switch Register: {broker_col}")
                    print(f"Switch Register Columns: {list(processed_df.columns)}")
                    
                    # Find transaction date column in Switch Register
                    tran_date_col = None
                    for col in processed_df.columns:
                        col_upper = str(col).upper().strip()
                        if 'TRAN' in col_upper and 'DATE' in col_upper:
                            tran_date_col = col
                            break
                        elif col_upper == 'TRANSACTION DATE' or col_upper == 'TRANSACTION_DATE':
                            tran_date_col = col
                            break
                        elif col_upper == 'DATE':
                            tran_date_col = col
                            break
                    
                    print(f"Transaction Date Column: {tran_date_col}")
                    
                    if cons_code_col and scheme_code_b_col and broker_col:
                        # Normalize Brokerage Structure for matching
                        brokerage_normalized = combined_brokerage_df.copy()
                        brokerage_normalized[cons_code_col] = brokerage_normalized[cons_code_col].astype(str).str.strip().str.upper()
                        brokerage_normalized[scheme_code_b_col] = brokerage_normalized[scheme_code_b_col].astype(str).str.strip().str.upper()
                        
                        # Remove NaN values that might cause issues
                        brokerage_normalized = brokerage_normalized[
                            (brokerage_normalized[cons_code_col] != 'NAN') &
                            (brokerage_normalized[cons_code_col] != '') &
                            (brokerage_normalized[scheme_code_b_col] != 'NAN') &
                            (brokerage_normalized[scheme_code_b_col] != '')
                        ]
                        
                        # Convert Investment Period dates to datetime if columns exist
                        if investment_period_from_col and investment_period_from_col in brokerage_normalized.columns:
                            brokerage_normalized['_PERIOD_FROM_DT'] = pd.to_datetime(
                                brokerage_normalized[investment_period_from_col], 
                                errors='coerce',
                                dayfirst=True
                            )
                        else:
                            brokerage_normalized['_PERIOD_FROM_DT'] = pd.NaT
                        
                        if investment_period_to_col and investment_period_to_col in brokerage_normalized.columns:
                            brokerage_normalized['_PERIOD_TO_DT'] = pd.to_datetime(
                                brokerage_normalized[investment_period_to_col], 
                                errors='coerce',
                                dayfirst=True
                            )
                        else:
                            brokerage_normalized['_PERIOD_TO_DT'] = pd.NaT
                        
                        # Convert transaction date in processed_df if column exists
                        if tran_date_col and tran_date_col in processed_df.columns:
                            processed_df['_TRAN_DATE_DT'] = pd.to_datetime(
                                processed_df[tran_date_col], 
                                errors='coerce',
                                dayfirst=True
                            )
                        else:
                            processed_df['_TRAN_DATE_DT'] = pd.NaT
                        
                        # DEBUG: Print sample values
                        print(f"\n=== DEBUG: Sample Brokerage Values ===")
                        if len(brokerage_normalized) > 0:
                            print(f"Sample Cons Code values: {brokerage_normalized[cons_code_col].head(5).tolist()}")
                            print(f"Sample Scheme Code values: {brokerage_normalized[scheme_code_b_col].head(5).tolist()}")
                            if investment_period_from_col:
                                print(f"Sample Period From values: {brokerage_normalized[investment_period_from_col].head(3).tolist()}")
                            if investment_period_to_col:
                                print(f"Sample Period To values: {brokerage_normalized[investment_period_to_col].head(3).tolist()}")
                        
                        # Create a function to get trail rates
                        match_count = 0
                        no_match_count = 0
                        
                        def get_trail_rates_and_periods(row):
                            nonlocal match_count, no_match_count
                            try:
                                # Get broker code
                                broker = row.get(broker_col) if broker_col in row.index else None
                                if pd.isna(broker) or broker == '':
                                    no_match_count += 1
                                    return ['Not Found'] * 7  # 5 trail rates + 2 period columns
                                
                                broker_str = str(broker).strip().upper()
                                
                                # Get IN subfund code
                                in_subfund = row.get('IN subfund code') if 'IN subfund code' in row.index else None
                                if pd.isna(in_subfund) or in_subfund == 'Not Found' or in_subfund == '':
                                    no_match_count += 1
                                    return ['Not Found'] * 7
                                
                                in_subfund_str = str(in_subfund).strip().upper()
                                
                                # Get transaction date
                                tran_date = row.get('_TRAN_DATE_DT') if '_TRAN_DATE_DT' in row.index else None
                                
                                # DEBUG: Print first few attempts
                                if match_count + no_match_count < 5:
                                    print(f"\n=== DEBUG: Matching Attempt {match_count + no_match_count + 1} ===")
                                    print(f"Broker: '{broker_str}'")
                                    print(f"IN Subfund: '{in_subfund_str}'")
                                    print(f"Transaction Date: {tran_date}")
                                
                                # Filter brokerage structure: match Cons Code with broker AND Scheme Code with IN subfund code
                                mask = (
                                    (brokerage_normalized[cons_code_col] == broker_str) &
                                    (brokerage_normalized[scheme_code_b_col] == in_subfund_str)
                                )
                                matches = brokerage_normalized[mask]
                                
                                if matches.empty:
                                    no_match_count += 1
                                    if match_count + no_match_count <= 5:
                                        # Check partial matches for debugging
                                        broker_matches = brokerage_normalized[brokerage_normalized[cons_code_col] == broker_str]
                                        scheme_matches = brokerage_normalized[brokerage_normalized[scheme_code_b_col] == in_subfund_str]
                                        print(f"  No exact match found")
                                        print(f"  Broker-only matches: {len(broker_matches)}")
                                        print(f"  Scheme-only matches: {len(scheme_matches)}")
                                        if len(broker_matches) > 0:
                                            print(f"  Sample broker match Cons Codes: {broker_matches[cons_code_col].head(3).tolist()}")
                                        if len(scheme_matches) > 0:
                                            print(f"  Sample scheme match Scheme Codes: {scheme_matches[scheme_code_b_col].head(3).tolist()}")
                                    return ['Not Found'] * 7
                                
                                # Filter by date range if transaction date and period dates are available
                                if pd.notna(tran_date) and investment_period_from_col and investment_period_to_col:
                                    # Check each match to see if transaction date falls within the period
                                    date_filtered_matches = []
                                    for idx, match_row in matches.iterrows():
                                        period_from_dt = match_row.get('_PERIOD_FROM_DT')
                                        period_to_dt = match_row.get('_PERIOD_TO_DT')
                                        
                                        # If both period dates are valid, check if transaction date is within range
                                        if pd.notna(period_from_dt) and pd.notna(period_to_dt):
                                            if period_from_dt <= tran_date <= period_to_dt:
                                                date_filtered_matches.append(idx)
                                        # If period dates are missing, include the match (no date filtering)
                                        elif pd.isna(period_from_dt) or pd.isna(period_to_dt):
                                            date_filtered_matches.append(idx)
                                    
                                    if date_filtered_matches:
                                        # Use the first date-filtered match
                                        matches = matches.loc[date_filtered_matches]
                                        if match_count + no_match_count < 5:
                                            print(f"  ✓ Date-filtered match found! (Transaction date within period)")
                                    else:
                                        # No matches within date range
                                        no_match_count += 1
                                        if match_count + no_match_count <= 5:
                                            print(f"  ✗ No match within date range (Transaction date: {tran_date})")
                                            if len(matches) > 0:
                                                sample_match = matches.iloc[0]
                                                print(f"    Sample Period From: {sample_match.get('_PERIOD_FROM_DT')}")
                                                print(f"    Sample Period To: {sample_match.get('_PERIOD_TO_DT')}")
                                        return ['Not Found'] * 7
                                else:
                                    # No date filtering - use first match
                                    if match_count + no_match_count < 5:
                                        if pd.isna(tran_date):
                                            print(f"  ⚠ No transaction date - using first match without date filtering")
                                        else:
                                            print(f"  ⚠ No period dates in brokerage - using first match without date filtering")
                                
                                match_count += 1
                                if match_count <= 5:
                                    print(f"  ✓ Match found!")
                                
                                # Take first match
                                first_match = matches.iloc[0]
                                results = []
                                
                                # Get trail rates
                                for year in range(1, 6):
                                    if year in trail_rate_cols:
                                        trail_col = trail_rate_cols[year]
                                        if trail_col in first_match.index:
                                            rate_value = first_match[trail_col]
                                            if pd.notna(rate_value) and str(rate_value).strip() != '':
                                                results.append(rate_value)
                                            else:
                                                results.append('Not Found')
                                        else:
                                            results.append('Not Found')
                                    else:
                                        results.append('Not Found')
                                
                                # Get Investment Period From
                                if investment_period_from_col and investment_period_from_col in first_match.index:
                                    period_from = first_match[investment_period_from_col]
                                    if pd.notna(period_from) and str(period_from).strip() != '':
                                        results.append(period_from)
                                    else:
                                        results.append('Not Found')
                                else:
                                    results.append('Not Found')
                                
                                # Get Investment Period To
                                if investment_period_to_col and investment_period_to_col in first_match.index:
                                    period_to = first_match[investment_period_to_col]
                                    if pd.notna(period_to) and str(period_to).strip() != '':
                                        results.append(period_to)
                                    else:
                                        results.append('Not Found')
                                else:
                                    results.append('Not Found')
                                
                                return results
                            except Exception as e:
                                no_match_count += 1
                                if match_count + no_match_count <= 5:
                                    print(f"  Exception: {str(e)}")
                                return ['Not Found'] * 7
                        
                        # Apply trail rate and period extraction
                        trail_results = processed_df.apply(get_trail_rates_and_periods, axis=1, result_type='expand')
                        trail_results.columns = [f'switch in Trail Rate {i} year' for i in range(1, 6)] + ['Investment Period From', 'Investment Period To']
                        
                        # Add trail rate columns to processed_df
                        for year in range(1, 6):
                            col_name = f'switch in Trail Rate {year} year'
                            processed_df[col_name] = trail_results[col_name]
                        
                        # Add Investment Period columns
                        processed_df['Investment Period From'] = trail_results['Investment Period From']
                        processed_df['Investment Period To'] = trail_results['Investment Period To']
                        
                        # DEBUG: Print summary
                        print(f"\n=== DEBUG: Matching Summary (IN) ===")
                        print(f"Total matches: {match_count}")
                        print(f"Total no matches: {no_match_count}")
                        print(f"Total rows processed: {len(processed_df)}")
                        
                        # Now match with OUT subfund code for "switch out" trail rates
                        loading_window.update_status("Matching with Brokerage Structure (OUT)...")
                        
                        match_count_out = 0
                        no_match_count_out = 0
                        
                        def get_trail_rates_out(row):
                            nonlocal match_count_out, no_match_count_out
                            try:
                                # Get broker code
                                broker = row.get(broker_col) if broker_col in row.index else None
                                if pd.isna(broker) or broker == '':
                                    no_match_count_out += 1
                                    return ['Not Found'] * 5  # 5 trail rates only
                                
                                broker_str = str(broker).strip().upper()
                                
                                # Get OUT subfund code
                                out_subfund = row.get('out subfund code') if 'out subfund code' in row.index else None
                                if pd.isna(out_subfund) or out_subfund == 'Not Found' or out_subfund == '':
                                    no_match_count_out += 1
                                    return ['Not Found'] * 5
                                
                                out_subfund_str = str(out_subfund).strip().upper()
                                
                                # Get transaction date
                                tran_date = row.get('_TRAN_DATE_DT') if '_TRAN_DATE_DT' in row.index else None
                                
                                # Filter brokerage structure: match Cons Code with broker AND Scheme Code with OUT subfund code
                                mask = (
                                    (brokerage_normalized[cons_code_col] == broker_str) &
                                    (brokerage_normalized[scheme_code_b_col] == out_subfund_str)
                                )
                                matches = brokerage_normalized[mask]
                                
                                if matches.empty:
                                    no_match_count_out += 1
                                    return ['Not Found'] * 5
                                
                                # Filter by date range if transaction date and period dates are available
                                if pd.notna(tran_date) and investment_period_from_col and investment_period_to_col:
                                    # Check each match to see if transaction date falls within the period
                                    date_filtered_matches = []
                                    for idx, match_row in matches.iterrows():
                                        period_from_dt = match_row.get('_PERIOD_FROM_DT')
                                        period_to_dt = match_row.get('_PERIOD_TO_DT')
                                        
                                        # If both period dates are valid, check if transaction date is within range
                                        if pd.notna(period_from_dt) and pd.notna(period_to_dt):
                                            if period_from_dt <= tran_date <= period_to_dt:
                                                date_filtered_matches.append(idx)
                                        # If period dates are missing, include the match (no date filtering)
                                        elif pd.isna(period_from_dt) or pd.isna(period_to_dt):
                                            date_filtered_matches.append(idx)
                                    
                                    if date_filtered_matches:
                                        # Use the first date-filtered match
                                        matches = matches.loc[date_filtered_matches]
                                    else:
                                        # No matches within date range
                                        no_match_count_out += 1
                                        return ['Not Found'] * 5
                                
                                match_count_out += 1
                                
                                # Take first match
                                first_match = matches.iloc[0]
                                results = []
                                
                                # Get trail rates
                                for year in range(1, 6):
                                    if year in trail_rate_cols:
                                        trail_col = trail_rate_cols[year]
                                        if trail_col in first_match.index:
                                            rate_value = first_match[trail_col]
                                            if pd.notna(rate_value) and str(rate_value).strip() != '':
                                                results.append(rate_value)
                                            else:
                                                results.append('Not Found')
                                        else:
                                            results.append('Not Found')
                                    else:
                                        results.append('Not Found')
                                
                                return results
                            except Exception as e:
                                no_match_count_out += 1
                                return ['Not Found'] * 5
                        
                        # Apply trail rate extraction for OUT subfund code
                        trail_results_out = processed_df.apply(get_trail_rates_out, axis=1, result_type='expand')
                        trail_results_out.columns = [f'switch out Trail Rate {i} year' for i in range(1, 6)]
                        
                        # Add "switch out" trail rate columns to processed_df
                        for year in range(1, 6):
                            col_name = f'switch out Trail Rate {year} year'
                            processed_df[col_name] = trail_results_out[col_name]
                        
                        # DEBUG: Print summary for OUT
                        print(f"\n=== DEBUG: Matching Summary (OUT) ===")
                        print(f"Total matches: {match_count_out}")
                        print(f"Total no matches: {no_match_count_out}")
                        
                        # Add check columns: if switch in > switch out, show "Check"
                        loading_window.update_status("Calculating checks...")
                        
                        def compare_trail_rates(value_in, value_out):
                            """Compare two trail rate values and return 'Check' if in > out"""
                            try:
                                # Handle "Not Found" or empty values
                                if pd.isna(value_in) or str(value_in).strip() == '' or str(value_in).strip().upper() == 'NOT FOUND':
                                    return ''
                                if pd.isna(value_out) or str(value_out).strip() == '' or str(value_out).strip().upper() == 'NOT FOUND':
                                    return ''
                                
                                # Convert to numeric, handling percentage signs and other formats
                                val_in_str = str(value_in).strip().replace('%', '').replace(',', '')
                                val_out_str = str(value_out).strip().replace('%', '').replace(',', '')
                                
                                try:
                                    val_in_num = float(val_in_str)
                                    val_out_num = float(val_out_str)
                                    
                                    # If switch in > switch out, return "Check"
                                    if val_in_num > val_out_num:
                                        return 'Check'
                                    else:
                                        return ''
                                except (ValueError, TypeError):
                                    # If conversion fails, return empty
                                    return ''
                            except Exception:
                                return ''
                        
                        # Add check columns for each year
                        for year in range(1, 6):
                            col_in = f'switch in Trail Rate {year} year'
                            col_out = f'switch out Trail Rate {year} year'
                            check_col = f'Check {year} year'
                            
                            if col_in in processed_df.columns and col_out in processed_df.columns:
                                processed_df[check_col] = processed_df.apply(
                                    lambda row: compare_trail_rates(row[col_in], row[col_out]),
                                    axis=1
                                )
                            else:
                                processed_df[check_col] = ''
                    else:
                        # Missing required columns
                        missing_cols = []
                        if not cons_code_col:
                            missing_cols.append("Cons Code")
                        if not scheme_code_b_col:
                            missing_cols.append("Scheme Code")
                        if not broker_col:
                            missing_cols.append("Broker Column (BROK_DLR_N or similar)")
                        
                        self.root.after(0, lambda m=missing_cols: messagebox.showwarning(
                            "Warning",
                            f"Columns not found:\n"
                            f"Brokerage Structure: {', '.join([c for c in m if c != 'Broker Column (BROK_DLR_N or similar)'])}\n"
                            f"Switch Register: {', '.join([c for c in m if c == 'Broker Column (BROK_DLR_N or similar)'])}\n"
                            f"Trail rates will not be added."
                        ))
                        # Add empty trail rate columns
                        for year in range(1, 6):
                            processed_df[f'switch in Trail Rate {year} year'] = 'Not Found'
                            processed_df[f'switch out Trail Rate {year} year'] = 'Not Found'
                            processed_df[f'Check {year} year'] = ''
                        processed_df['Investment Period From'] = 'Not Found'
                        processed_df['Investment Period To'] = 'Not Found'
                else:
                    # No brokerage structure data
                    for year in range(1, 6):
                        processed_df[f'switch in Trail Rate {year} year'] = 'Not Found'
                        processed_df[f'switch out Trail Rate {year} year'] = 'Not Found'
                        processed_df[f'Check {year} year'] = ''
                    processed_df['Investment Period From'] = 'Not Found'
                    processed_df['Investment Period To'] = 'Not Found'
                
                # Reorder columns: Group by year (switch in, switch out, check for each year)
                loading_window.update_status("Reordering columns...")
                
                # Get all current columns
                all_columns = list(processed_df.columns)
                
                # Find columns to reorder
                trail_rate_columns = {}
                check_columns = {}
                other_columns = []
                
                for col in all_columns:
                    try:
                        # Ensure col is a string before using 'in' operator
                        col_str = str(col) if not isinstance(col, str) else col
                        
                        if 'switch in Trail Rate' in col_str:
                            # Extract year number
                            year_match = None
                            for year in range(1, 6):
                                if f'{year} year' in col_str:
                                    year_match = year
                                    break
                            if year_match:
                                if year_match not in trail_rate_columns:
                                    trail_rate_columns[year_match] = {}
                                trail_rate_columns[year_match]['in'] = col
                        elif 'switch out Trail Rate' in col_str:
                            # Extract year number
                            year_match = None
                            for year in range(1, 6):
                                if f'{year} year' in col_str:
                                    year_match = year
                                    break
                            if year_match:
                                if year_match not in trail_rate_columns:
                                    trail_rate_columns[year_match] = {}
                                trail_rate_columns[year_match]['out'] = col
                        elif 'Check' in col_str and 'year' in col_str:
                            # Extract year number
                            year_match = None
                            for year in range(1, 6):
                                if f'{year} year' in col_str:
                                    year_match = year
                                    break
                            if year_match:
                                check_columns[year_match] = col
                        else:
                            other_columns.append(col)
                    except (TypeError, AttributeError):
                        # Skip columns that can't be converted to string
                        other_columns.append(col)
                
                # Build new column order: group by year
                new_column_order = []
                
                # Add other columns first (before trail rate columns)
                trail_rate_start_idx = None
                for idx, col in enumerate(all_columns):
                    try:
                        # Ensure col is a string before using 'in' operator
                        col_str = str(col) if not isinstance(col, str) else col
                        if 'switch in Trail Rate' in col_str or 'switch out Trail Rate' in col_str or ('Check' in col_str and 'year' in col_str):
                            if trail_rate_start_idx is None:
                                trail_rate_start_idx = idx
                            break
                        new_column_order.append(col)
                    except (TypeError, AttributeError):
                        new_column_order.append(col)
                
                # Add trail rate columns grouped by year
                for year in range(1, 6):
                    if year in trail_rate_columns:
                        if 'in' in trail_rate_columns[year]:
                            new_column_order.append(trail_rate_columns[year]['in'])
                        if 'out' in trail_rate_columns[year]:
                            new_column_order.append(trail_rate_columns[year]['out'])
                    if year in check_columns:
                        new_column_order.append(check_columns[year])
                
                # Add remaining columns that weren't in the original order
                for col in all_columns:
                    if col not in new_column_order:
                        new_column_order.append(col)
                
                # Reorder the dataframe
                processed_df = processed_df[new_column_order]
                
                # Add check for Regular vs Direct scheme matching
                loading_window.update_status("Checking Regular vs Direct scheme matches...")
                
                def check_regular_direct_match(row):
                    """Check if switch in and switch out schemes are the same but one is Regular and other is Direct"""
                    try:
                        switch_in = row.get('switch in scheme') if 'switch in scheme' in row.index else None
                        switch_out = row.get('switch out scheme') if 'switch out scheme' in row.index else None
                        
                        if pd.isna(switch_in) or pd.isna(switch_out):
                            return ''
                        
                        # Ensure we have strings
                        try:
                            switch_in_str = str(switch_in).strip() if switch_in is not None else ''
                            switch_out_str = str(switch_out).strip() if switch_out is not None else ''
                        except Exception:
                            return ''
                        
                        if switch_in_str == '' or switch_out_str == '':
                            return ''
                        
                        # Remove code prefix (e.g., "129B/ABSL" or any code before "/")
                        def remove_code_prefix(scheme_str):
                            try:
                                scheme_str = str(scheme_str) if not isinstance(scheme_str, str) else scheme_str
                                if '/' in scheme_str:
                                    # Split by "/" and take everything after the first "/"
                                    parts = scheme_str.split('/', 1)
                                    if len(parts) > 1:
                                        return parts[1].strip()
                                return scheme_str
                            except Exception:
                                return str(scheme_str) if scheme_str is not None else ''
                        
                        scheme_in = remove_code_prefix(switch_in_str)
                        scheme_out = remove_code_prefix(switch_out_str)
                        
                        # Normalize by removing plan type indicators (Regular, Direct, Reg, etc.)
                        def normalize_scheme_name(scheme):
                            try:
                                # Ensure scheme is a string
                                scheme = str(scheme) if not isinstance(scheme, str) else scheme
                                # Remove common plan type indicators
                                scheme_upper = scheme.upper()
                                # Remove: Regular Growth, Reg Growth, Regular, Reg Plan, Reg, etc.
                                scheme_upper = scheme_upper.replace('REGULAR GROWTH', '').replace('REG GROWTH', '')
                                scheme_upper = scheme_upper.replace('REGULAR', '').replace('REG PLAN', '')
                                scheme_upper = scheme_upper.replace('REG', '').replace('GROWTH', '')
                                # Remove: Direct Growth, Direct, etc.
                                scheme_upper = scheme_upper.replace('DIRECT GROWTH', '').replace('DIRECT', '')
                                # Remove extra spaces and dashes
                                scheme_upper = scheme_upper.replace('-', '').strip()
                                # Remove multiple spaces
                                while '  ' in scheme_upper:
                                    scheme_upper = scheme_upper.replace('  ', ' ')
                                return scheme_upper.strip()
                            except Exception:
                                return str(scheme) if scheme is not None else ''
                        
                        normalized_in = normalize_scheme_name(scheme_in)
                        normalized_out = normalize_scheme_name(scheme_out)
                        
                        # Check if normalized scheme names match
                        if normalized_in != normalized_out:
                            return ''
                        
                        # Check if one has Regular and other has Direct
                        try:
                            # Ensure we have strings before calling .upper()
                            switch_in_upper = str(switch_in_str).upper() if switch_in_str is not None else ''
                            switch_out_upper = str(switch_out_str).upper() if switch_out_str is not None else ''
                            
                            # Additional safety check
                            if not isinstance(switch_in_upper, str) or not isinstance(switch_out_upper, str):
                                return ''
                            
                            # Check for Regular indicators
                            has_regular_in = any(indicator in switch_in_upper for indicator in ['REGULAR', 'REG GROWTH', 'REG PLAN', 'REG '])
                            has_regular_out = any(indicator in switch_out_upper for indicator in ['REGULAR', 'REG GROWTH', 'REG PLAN', 'REG '])
                            
                            # Check for Direct indicators
                            has_direct_in = 'DIRECT' in switch_in_upper
                            has_direct_out = 'DIRECT' in switch_out_upper
                        except (TypeError, AttributeError):
                            return ''
                        
                        # If one is Regular and other is Direct, show Check
                        if (has_regular_in and has_direct_out) or (has_direct_in and has_regular_out):
                            return 'Check'
                        
                        return ''
                    except Exception as e:
                        return ''
                
                # Add the check column
                if 'switch in scheme' in processed_df.columns and 'switch out scheme' in processed_df.columns:
                    processed_df['Regular vs Direct Check'] = processed_df.apply(check_regular_direct_match, axis=1)
                else:
                    processed_df['Regular vs Direct Check'] = ''
                
                loading_window.update_status("Processing complete...")
                time.sleep(0.3)
                
                # Remove rows where broker is "DIRECT"
                loading_window.update_status("Filtering out DIRECT broker entries...")
                
                # Find broker column (it should already be found, but check again to be safe)
                broker_col_filter = None
                for col in processed_df.columns:
                    col_upper = str(col).upper().strip()
                    if 'BROK' in col_upper and ('DLR' in col_upper or 'DEALER' in col_upper):
                        broker_col_filter = col
                        break
                    elif col_upper == 'BROKER' or col_upper == 'BROKER CODE' or col_upper == 'BROKER_CODE':
                        broker_col_filter = col
                        break
                
                if broker_col_filter and broker_col_filter in processed_df.columns:
                    initial_count = len(processed_df)
                    # Filter out rows where broker is "DIRECT" (case-insensitive)
                    processed_df = processed_df[
                        ~processed_df[broker_col_filter].astype(str).str.strip().str.upper().eq('DIRECT')
                    ]
                    removed_count = initial_count - len(processed_df)
                    if removed_count > 0:
                        print(f"\n=== Removed {removed_count} rows with DIRECT broker ===")
                        self.root.after(0, lambda: messagebox.showinfo(
                            "Filter Applied",
                            f"Removed {removed_count} row(s) where broker is 'DIRECT'.\n"
                            f"Remaining rows: {len(processed_df)}"
                        ))
                else:
                    print("\n=== Warning: Broker column not found, skipping DIRECT filter ===")
                
                loading_window.update_status("Preparing to save...")
                save_path = filedialog.asksaveasfilename(
                    title="Save Processed File",
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
                )
                
                if save_path:
                    try:
                        loading_window.update_status("Saving file...")
                        # Save processed data to Excel
                        wb = Workbook()
                        ws = wb.active
                        ws.title = 'Processed Data'
                        
                        # Write headers
                        for col_idx, col_name in enumerate(processed_df.columns, 1):
                            cell = ws.cell(row=1, column=col_idx, value=col_name)
                            cell.fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
                            cell.font = Font(color='FFFFFF', bold=True, size=11, name='Calibri')
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        
                        # Write data rows
                        for row_idx, row in enumerate(processed_df.itertuples(index=False), 2):
                            for col_idx, value in enumerate(row, 1):
                                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                                cell.font = Font(size=10, name='Calibri')
                                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        
                        # Auto-adjust column widths
                        for col_idx in range(1, ws.max_column + 1):
                            col_letter = get_column_letter(col_idx)
                            max_length = 0
                            for row_idx in range(1, ws.max_row + 1):
                                cell = ws.cell(row=row_idx, column=col_idx)
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            adjusted_width = min(max(max_length + 2, 12), 50)
                            ws.column_dimensions[col_letter].width = adjusted_width
                        
                        wb.save(save_path)
                        
                        loading_window.update_status("Complete!")
                        time.sleep(0.5)
                        
                        self.root.after(0, lambda: messagebox.showinfo(
                            "Success",
                            f"File processed and saved successfully at:\n{save_path}"
                        ))
                        self.root.after(0, lambda: self.status_label.configure(
                            text="Processing completed successfully!",
                            text_color="#27ae60"
                        ))
                    except Exception as e:
                        self.root.after(0, lambda: messagebox.showerror(
                            "Error",
                            f"Could not save file:\n{e}"
                        ))
                        self.root.after(0, lambda: self.status_label.configure(
                            text="Status: Could not save file!",
                            text_color="#e74c3c"
                        ))
                else:
                    self.root.after(0, lambda: messagebox.showinfo(
                        "Cancelled",
                        "File save was cancelled."
                    ))
                    self.root.after(0, lambda: self.status_label.configure(
                        text="Status: File save was cancelled.",
                        text_color="#f39c12"
                    ))
                
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror(
                    "Error",
                    f"Error processing files: {str(e)}"
                ))
                self.root.after(0, lambda: self.status_label.configure(
                    text=f"Status: Error processing files: {str(e)}",
                    text_color="#e74c3c"
                ))
            finally:
                self.root.after(0, loading_window.stop)
        
        # Run processing in a separate thread
        thread = threading.Thread(target=process_file)
        thread.daemon = True
        thread.start()


def main():
    root = ctk.CTk()
    app = SwitchRegisterGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()

